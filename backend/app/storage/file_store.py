"""Handles writing/reading uploaded files and their Parquet derivatives."""
import hashlib
from pathlib import Path
import polars as pl
import openpyxl
from app.config import settings


def _storage_root() -> Path:
    return Path(settings.file_storage_path)


def parquet_path(file_id: str, sheet_name: str) -> Path:
    safe = sheet_name.replace("/", "_").replace("\\", "_")
    return _storage_root() / file_id / f"{safe}.parquet"


def raw_path(file_id: str, filename: str) -> Path:
    return _storage_root() / file_id / filename


def save_raw(file_id: str, filename: str, data: bytes) -> Path:
    dest = raw_path(file_id, filename)
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)
    return dest


def parse_excel(file_id: str, raw: Path) -> dict[str, dict]:
    """Parse all sheets → Parquet. Returns sheet metadata."""
    wb = openpyxl.load_workbook(raw, read_only=True, data_only=True)
    sheets: dict[str, dict] = {}

    for name in wb.sheetnames:
        ws = wb[name]
        # Read rows; first row = headers
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[name] = {"row_count": 0, "columns": []}
            continue

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        df = pl.DataFrame(
            {h: [r[i] for r in data_rows] for i, h in enumerate(headers)},
            strict=False,
            infer_schema_length=500,
        )

        dest = parquet_path(file_id, name)
        dest.parent.mkdir(parents=True, exist_ok=True)
        df.write_parquet(dest)

        sheets[name] = _sheet_meta(name, df)

    wb.close()
    return sheets


def parse_csv(file_id: str, raw: Path) -> dict[str, dict]:
    """Parse CSV → Parquet. Returns sheet metadata keyed as 'Sheet1'."""
    df = pl.read_csv(raw, infer_schema_length=500, ignore_errors=True)
    dest = parquet_path(file_id, "Sheet1")
    dest.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(dest)
    return {"Sheet1": _sheet_meta("Sheet1", df)}


def read_sheet(file_id: str, sheet_name: str) -> pl.DataFrame:
    path = parquet_path(file_id, sheet_name)
    if not path.exists():
        raise FileNotFoundError(f"Sheet not found: {sheet_name}")
    return pl.read_parquet(path)


def _sheet_meta(name: str, df: pl.DataFrame) -> dict:
    return {
        "name": name,
        "row_count": len(df),
        "columns": [_col_meta(col, df[col]) for col in df.columns],
    }


def _col_meta(name: str, series: pl.Series) -> dict:
    dtype = series.dtype
    if dtype in (pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64):
        col_type = "integer"
    elif dtype in (pl.Float32, pl.Float64):
        col_type = "number"
    elif dtype == pl.Boolean:
        col_type = "boolean"
    elif dtype == pl.Date or dtype == pl.Datetime:
        col_type = "date"
    else:
        col_type = "string"

    non_null = series.drop_nulls()
    samples = [str(v) for v in non_null.head(5).to_list()]

    return {
        "name": name,
        "type": col_type,
        "nullable": series.null_count() > 0,
        "sample_values": samples,
    }


def md5(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()
